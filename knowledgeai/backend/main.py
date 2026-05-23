"""
main.py — FastAPI application: ingestion, query, document management, analytics.
"""
import asyncio
import json
import logging
import os
import shutil
import tempfile
import time
from typing import List, Optional

from fastapi import (FastAPI, File, Form, HTTPException,
                     UploadFile, WebSocket, WebSocketDisconnect)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

# ── internal imports ──────────────────────────────────────────────────────────
from backend.config import BRD_FOLDER, UPLOADS_DIR, HISTORY_ENABLED, EXCEL_OUTPUT_DIR, IMAGES_DIR, LLM_MODEL
from backend.ingestion import metadata_db as db
from backend.ingestion.orchestrator import enqueue, worker
from backend.ingestion.embedder import embed_single
from backend.retrieval.hybrid_search import hybrid_search
from backend.retrieval.reranker import rerank, is_confident
from backend.generation.cache import cache_lookup, cache_store, cache_stats
from backend.generation.prompt_builder import build_prompt, build_fallback_message
from backend.generation.llm import generate, stream_generate
from backend.generation.formatter import format_response
from backend.observability.logger import setup_logging, QueryMetrics

setup_logging()
logger = logging.getLogger(__name__)

# ── App ───────────────────────────────────────────────────────────────────────
app = FastAPI(title="KnowledgeAI", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve the single-file frontend
FRONTEND_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend")

# Serve extracted document images as static files
app.mount("/static/images", StaticFiles(directory=IMAGES_DIR), name="doc_images")


@app.on_event("startup")
async def startup():
    db.init_db()
    asyncio.create_task(worker())
    # Start the file watcher using THIS running event loop (not a separate one)
    loop = asyncio.get_event_loop()
    try:
        from backend.ingestion.watcher import start_watcher
        start_watcher(loop)
    except Exception as e:
        logger.warning("File watcher could not start (non-critical): %s", e)
    # Auto-ingest BRDs folder — slight delay so worker task is ready
    asyncio.create_task(_ingest_brd_folder())
    logger.info("KnowledgeAI server started — BRD folder: %s", BRD_FOLDER)


async def _ingest_brd_folder():
    """Scan BRDs folder and enqueue any supported files."""
    await asyncio.sleep(0.5)   # let the worker task spin up first
    supported = {".pdf", ".docx", ".doc", ".txt", ".md"}
    brd_path = os.path.abspath(BRD_FOLDER)
    if not os.path.isdir(brd_path):
        logger.warning("BRDs folder not found at: %s", brd_path)
        return
    files = []
    for root, _, filenames in os.walk(brd_path):
        for f in filenames:
            if os.path.splitext(f)[1].lower() in supported:
                files.append(os.path.join(root, f))
    logger.info("BRDs folder scan: found %d file(s) in %s", len(files), brd_path)
    for full_path in files:
        # Use relative path from BRDs root as the display name (preserves subfolder context)
        rel_name = os.path.relpath(full_path, brd_path)
        logger.info("  Queuing BRD file: %s", rel_name)
        await enqueue(full_path, folder="BRDs", original_name=rel_name)


# ── Schemas ───────────────────────────────────────────────────────────────────
ALLOWED_MODELS = {
    # OpenAI
    "gpt-4o", "gpt-4o-mini", "gpt-4-turbo", "gpt-3.5-turbo",
    # Google Gemini
    "gemini-2.0-flash", "gemini-1.5-pro", "gemini-1.5-flash",
    # Anthropic
    "claude-opus-4-6", "claude-sonnet-4-6", "claude-haiku-4-5-20251001",
}

class QueryRequest(BaseModel):
    query: str
    mode: str = "balanced"   # precise | balanced | exploratory
    folder: Optional[str] = None
    model: Optional[str] = None  # override LLM model per query


class FeedbackRequest(BaseModel):
    query_id: int
    vote: int  # 1 = thumbs up, -1 = thumbs down

class HistoryFeedbackRequest(BaseModel):
    history_id: int
    vote: int

class PromoteFAQRequest(BaseModel):
    history_id: int

class UpdateFAQRequest(BaseModel):
    answer: str


# ── Image helper ──────────────────────────────────────────────────────────────
def _images_for_chunks(chunks: list) -> list:
    """Return image payloads for the given retrieved chunks."""
    raw = db.get_images_for_chunks(chunks)
    return [
        {
            "doc_id":    r["doc_id"],
            "file_name": r["file_name"],
            "page_num":  r["page_num"],
            "url":       f"/static/images/{r['image_filename']}",
        }
        for r in raw
    ]


# ── Shared query logic ────────────────────────────────────────────────────────
async def _run_query(query: str, mode: str = "balanced",
                     source_type: str = "ui", model: str = None) -> dict:
    """Core RAG pipeline. Returns full response payload + history_id."""
    metrics = QueryMetrics()

    # 1. FAQ exact match (instant, no LLM)
    faq = await asyncio.to_thread(db.faq_lookup, query)
    if faq:
        payload = {
            "answer": faq["answer"],
            "sources": [{"file_name": s, "score": 1.0} for s in (faq["sources"] or [])],
            "confidence": 1.0,
            "cache_hit": True,
            "follow_ups": [],
            "faq_hit": True,
        }
        history_id = None
        if HISTORY_ENABLED:
            history_id = await asyncio.to_thread(
                db.save_history, query, faq["answer"],
                faq["sources"] or [], 1.0, True, 0, source_type
            )
        return {**payload, "latency_ms": metrics.total_ms, "history_id": history_id}

    # 2. Embed
    query_vector = await asyncio.to_thread(embed_single, query)
    metrics.mark("embed")

    # 3. Semantic cache
    cached = cache_lookup(query_vector, model=model or LLM_MODEL)
    if cached:
        metrics.mark("cache_hit")
        db.log_query(query, len(cached["answer"]), len(cached["sources"]),
                     cached["confidence"], True, metrics.total_ms)
        history_id = None
        if HISTORY_ENABLED:
            src_names = [s["file_name"] for s in cached.get("sources", [])]
            history_id = await asyncio.to_thread(
                db.save_history, query, cached["answer"],
                src_names, cached["confidence"], True, metrics.total_ms, source_type
            )
        return {**cached, "latency_ms": metrics.total_ms, "history_id": history_id}

    # 4. Hybrid search
    hits = await asyncio.to_thread(hybrid_search, query_vector, query)
    metrics.mark("retrieval")

    # 5. Re-rank
    top_chunks, confidence = rerank(query, hits)
    metrics.mark("rerank")

    # 6. Confidence threshold
    if not is_confident(confidence) or not top_chunks:
        answer  = build_fallback_message(query)
        payload = format_response(answer, [], confidence, False)
        db.log_query(query, len(answer), 0, confidence, False, metrics.total_ms)
        history_id = None
        if HISTORY_ENABLED:
            history_id = await asyncio.to_thread(
                db.save_history, query, answer, [], confidence, False,
                metrics.total_ms, source_type
            )
        return {**payload, "latency_ms": metrics.total_ms, "history_id": history_id}

    # 7. Build prompt + LLM
    messages = build_prompt(query, top_chunks, mode=mode)
    answer   = await generate(messages, model=model)
    metrics.mark("llm")

    # 8. Format + cache + log
    payload = format_response(answer, top_chunks, confidence, False)
    cache_store(query, query_vector, payload, model=model or LLM_MODEL)
    src_names = [s["file_name"] for s in payload["sources"]]
    db.log_query(query, len(answer), len(payload["sources"]),
                 confidence, False, metrics.total_ms)
    history_id = None
    if HISTORY_ENABLED:
        history_id = await asyncio.to_thread(
            db.save_history, query, answer, src_names,
            confidence, False, metrics.total_ms, source_type
        )

    # 9. Attach images (screenshots from source pages)
    images = await asyncio.to_thread(_images_for_chunks, top_chunks)

    return {**payload, "latency_ms": metrics.total_ms,
            "history_id": history_id, "images": images}


# ── Query endpoint ────────────────────────────────────────────────────────────
@app.post("/api/query")
async def query_endpoint(req: QueryRequest):
    query = req.query.strip()
    if not query:
        raise HTTPException(400, "Query cannot be empty")
    model = req.model if req.model in ALLOWED_MODELS else None
    return await _run_query(query, mode=req.mode, source_type="ui", model=model)


# ── Streaming query (WebSocket) ───────────────────────────────────────────────
@app.websocket("/ws/query")
async def ws_query(websocket: WebSocket):
    await websocket.accept()
    try:
        data    = await websocket.receive_json()
        query   = data.get("query", "").strip()
        mode    = data.get("mode", "balanced")
        model   = data.get("model") if data.get("model") in ALLOWED_MODELS else None
        metrics = QueryMetrics()

        if not query:
            await websocket.send_json({"type": "error", "message": "Empty query"})
            return

        # Embed
        query_vector = await asyncio.to_thread(embed_single, query)

        # Cache check
        cached = cache_lookup(query_vector, model=model or LLM_MODEL)
        if cached:
            await websocket.send_json({"type": "cached", "data": cached})
            await websocket.send_json({"type": "done", "latency_ms": metrics.total_ms})
            return

        # Search + rerank
        hits        = await asyncio.to_thread(hybrid_search, query_vector, query)
        top_chunks, confidence = rerank(query, hits)

        # Send sources immediately (before tokens arrive)
        sources_payload = {
            "sources":    [{"file_name": c["file_name"],
                            "chunk_index": c.get("chunk_index", 0),
                            "score": round(c.get("rerank_score", 0), 3)}
                           for c in top_chunks],
            "confidence": confidence,
        }
        await websocket.send_json({"type": "sources", "data": sources_payload})

        # Send images associated with retrieved chunks
        images = await asyncio.to_thread(_images_for_chunks, top_chunks)
        if images:
            await websocket.send_json({"type": "images", "data": images})

        # Fallback
        if not is_confident(confidence) or not top_chunks:
            msg = build_fallback_message(query)
            await websocket.send_json({"type": "token", "token": msg})
            await websocket.send_json({"type": "done", "latency_ms": metrics.total_ms})
            return

        # Stream LLM tokens
        messages = build_prompt(query, top_chunks, mode=mode)
        full_answer = ""
        async for token in stream_generate(messages, model=model):
            full_answer += token
            await websocket.send_json({"type": "token", "token": token})

        # Cache + log
        payload = format_response(full_answer, top_chunks, confidence, False)
        cache_store(query, query_vector, payload, model=model or LLM_MODEL)
        db.log_query(query, len(full_answer), len(top_chunks),
                     confidence, False, metrics.total_ms)

        await websocket.send_json({"type": "done", "latency_ms": metrics.total_ms})

    except WebSocketDisconnect:
        logger.info("WS client disconnected")
    except Exception as e:
        logger.error("WS error: %s", e)
        try:
            await websocket.send_json({"type": "error", "message": str(e)})
        except Exception:
            pass


# ── Document upload ───────────────────────────────────────────────────────────
@app.post("/api/upload")
async def upload_documents(
    files: List[UploadFile] = File(...),
    folder: str = Form("BRDs"),
):
    results = []
    for uf in files:
        suffix = os.path.splitext(uf.filename)[1]
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            content = await uf.read()
            tmp.write(content)
            tmp_path = tmp.name

        await enqueue(tmp_path, folder=folder, original_name=uf.filename)
        results.append({"file": uf.filename, "status": "queued"})

    return {"uploaded": results}


# ── Trigger BRD folder re-scan ────────────────────────────────────────────────
@app.post("/api/ingest/folder")
async def ingest_folder():
    await _ingest_brd_folder()
    return {"status": "queued", "folder": BRD_FOLDER}


# ── Document listing & management ─────────────────────────────────────────────
@app.get("/api/documents")
async def list_documents(folder: Optional[str] = None):
    docs = await asyncio.to_thread(db.list_documents, folder)
    return {"documents": docs}


@app.delete("/api/documents/{doc_id}")
async def delete_document(doc_id: str):
    from backend.retrieval.vector_store import delete_doc_vectors
    doc = db.get_document(doc_id)
    if not doc:
        raise HTTPException(404, "Document not found")
    db.delete_document(doc_id)
    await asyncio.to_thread(delete_doc_vectors, doc_id, doc.get("version_id", 1))
    return {"status": "deleted", "doc_id": doc_id}


# ── Analytics ─────────────────────────────────────────────────────────────────
@app.get("/api/analytics")
async def analytics():
    data  = await asyncio.to_thread(db.get_analytics)
    cache = cache_stats()
    return {**data, "cache": cache}


# ── Feedback ──────────────────────────────────────────────────────────────────
@app.post("/api/feedback")
async def feedback(req: FeedbackRequest):
    db.record_feedback(req.query_id, req.vote)
    return {"status": "recorded"}


@app.post("/api/history/feedback")
async def history_feedback(req: HistoryFeedbackRequest):
    await asyncio.to_thread(db.update_history_feedback, req.history_id, req.vote)
    return {"status": "recorded"}


# ── Q&A History ───────────────────────────────────────────────────────────────
@app.get("/api/history")
async def get_history(limit: int = 100, offset: int = 0,
                      source_type: Optional[str] = None):
    rows = await asyncio.to_thread(db.get_history, limit, offset, source_type)
    return {"history": rows, "history_enabled": HISTORY_ENABLED}


@app.get("/api/history/enabled")
async def history_status():
    return {"history_enabled": HISTORY_ENABLED}


# ── FAQ Management ────────────────────────────────────────────────────────────
@app.get("/api/faq")
async def list_faqs():
    rows = await asyncio.to_thread(db.get_faqs)
    return {"faqs": rows}


@app.post("/api/faq/promote")
async def promote_faq(req: PromoteFAQRequest):
    faq_id = await asyncio.to_thread(db.promote_to_faq, req.history_id)
    return {"status": "promoted", "faq_id": faq_id}


@app.put("/api/faq/{faq_id}")
async def update_faq(faq_id: int, req: UpdateFAQRequest):
    await asyncio.to_thread(db.update_faq, faq_id, req.answer)
    return {"status": "updated"}


@app.delete("/api/faq/{faq_id}")
async def delete_faq(faq_id: int):
    await asyncio.to_thread(db.delete_faq, faq_id)
    return {"status": "deleted"}


# ── Excel Batch Q&A ───────────────────────────────────────────────────────────
@app.post("/api/excel-qa")
async def excel_qa(
    file: UploadFile = File(...),
    question_col: str = Form("B"),      # Excel column letter containing questions
    answer_col:   str = Form("E"),      # Column to write answers into
    mode:         str = Form("balanced"),
    sheet_name:   str = Form("0"),      # Sheet index (0) or name
):
    """
    Upload an Excel file with questions. Answers are written to answer_col.
    The updated file is saved to EXCEL_OUTPUT_DIR and also returned for download.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import column_index_from_string
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    # Save upload to temp file
    suffix = os.path.splitext(file.filename)[1] or ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        wb = load_workbook(tmp_path)
        sheet_idx = int(sheet_name) if sheet_name.isdigit() else None
        ws = wb.worksheets[sheet_idx] if sheet_idx is not None else wb[sheet_name]

        q_col = column_index_from_string(question_col.upper())
        a_col = column_index_from_string(answer_col.upper())

        # Write header for answer column
        ws.cell(row=1, column=a_col).value = "KnowledgeAI Answer"
        ws.cell(row=1, column=a_col).font = Font(bold=True)
        ws.cell(row=1, column=a_col).fill = PatternFill("solid", start_color="FFF2CC")

        results = []
        total_rows = ws.max_row

        for row_idx in range(2, total_rows + 1):
            cell = ws.cell(row=row_idx, column=q_col)
            question = str(cell.value or "").strip()
            # Skip blank rows and section headers (emoji/short text)
            if not question or len(question) < 10 or question.startswith("🔹"):
                continue

            try:
                result = await _run_query(question, mode=mode, source_type="excel")
                answer = result.get("answer", "")
            except Exception as e:
                answer = f"[Error: {e}]"

            ans_cell = ws.cell(row=row_idx, column=a_col)
            ans_cell.value = answer
            ans_cell.alignment = Alignment(wrap_text=True, vertical="top")
            results.append({"row": row_idx, "question": question[:80], "answered": bool(answer)})

        # Set column width
        ws.column_dimensions[answer_col.upper()].width = 80

        # Save output
        base_name = os.path.splitext(file.filename)[0]
        out_name  = f"{base_name}_KnowledgeAI_Answers.xlsx"
        out_path  = os.path.join(EXCEL_OUTPUT_DIR, out_name)
        wb.save(out_path)

        return {
            "status": "done",
            "output_file": out_name,
            "output_path": out_path,
            "rows_processed": len(results),
            "rows": results,
        }

    finally:
        os.unlink(tmp_path)


# ── Health ────────────────────────────────────────────────────────────────────
@app.get("/api/health")
async def health():
    from backend.retrieval.vector_store import collection_count
    return {
        "status": "ok",
        "vector_count": collection_count(),
        "cache": cache_stats(),
    }


# ── Ingestion status (debug) ───────────────────────────────────────────────────
@app.get("/api/ingest/status")
async def ingest_status():
    """Shows what files are in the BRDs folder vs what's indexed."""
    supported = {".pdf", ".docx", ".doc", ".txt", ".md"}
    brd_path = os.path.abspath(BRD_FOLDER)
    on_disk = []
    if os.path.isdir(brd_path):
        for root, _, filenames in os.walk(brd_path):
            for f in filenames:
                if os.path.splitext(f)[1].lower() in supported:
                    on_disk.append(os.path.relpath(os.path.join(root, f), brd_path))
    indexed = await asyncio.to_thread(db.list_documents)
    recent_log = await asyncio.to_thread(_get_recent_ingestion_log)
    return {
        "brd_folder": brd_path,
        "files_on_disk": on_disk,
        "files_indexed": [d["file_name"] for d in indexed],
        "recent_ingestion_log": recent_log,
    }

def _get_recent_ingestion_log():
    import sqlite3
    from backend.config import SQLITE_PATH
    con = sqlite3.connect(SQLITE_PATH)
    con.row_factory = sqlite3.Row
    rows = con.execute(
        "SELECT * FROM ingestion_log ORDER BY created_at DESC LIMIT 30"
    ).fetchall()
    con.close()
    return [dict(r) for r in rows]


# ── Serve frontend ────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def serve_ui():
    index = os.path.join(FRONTEND_DIR, "index.html")
    if os.path.exists(index):
        with open(index, "r", encoding="utf-8") as f:
            return HTMLResponse(f.read())
    return HTMLResponse("<h1>KnowledgeAI — frontend not found</h1>")
